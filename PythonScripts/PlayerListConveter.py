# bad code demo but it surely gets the job done
# this may not work for other versions of this spreadsheet but who cares

# TODO: Add openpyxl to requirements
# TODO: ConvertNormalPlayers function
from openpyxl import load_workbook
import json

# Haland would be the at 4th row, and Nick Pope, at the 110th (last)
# Only converts Legendary and Rare players
def playerConverter(ws: str, fileName: str):
    workBook = load_workbook(fileName)
    workingSheet = workBook.get_sheet_by_name(ws)

    playerDict = dict()
    for i in range(4, 110):
        # fucking boilerplate shit
        plName = workingSheet.cell(i, 1).value
        plSurname = workingSheet.cell(i, 2).value

        plPrice = workingSheet.cell(i, 3).value.__int__()
        plCountry = workingSheet.cell(i, 4).value

        plPos = workingSheet.cell(i, 6).value
        plRating = workingSheet.cell(i, 8).value.__int__()
        plFoot = workingSheet.cell(i, 7).value
        plClub = workingSheet.cell(i, 5).value
        plHeight = workingSheet.cell(i, 9).value.__int__()

        plSpeed = workingSheet.cell(i, 10).value.__int__()
        plAcc = workingSheet.cell(i, 11).value.__int__()
        plStamina = int(workingSheet.cell(i, 12).value) if workingSheet.cell(i, 12).value is not None else None
        plControl = workingSheet.cell(i, 13).value.__int__()
        plStrength = workingSheet.cell(i, 14).value.__int__()
        plTackle = workingSheet.cell(i, 15).value.__int__()
        plPassing = workingSheet.cell(i, 16).value.__int__()
        plShooting = int(workingSheet.cell(i, 17).value) if workingSheet.cell(i, 17).value is not None else None
        plGKHandling = int(workingSheet.cell(i, 18).value) if workingSheet.cell(i, 18).value is not None else None
        plGKReaction = int(workingSheet.cell(i, 19).value) if workingSheet.cell(i, 19).value is not None else None

        if plName == None: plFullname = plSurname.lower()
        else: plFullname = f'{plName.lower()} {plSurname.lower()}'
        playerDict[plFullname] = {
            'gameInfo': {
                'name': plName, 
                'surname': plSurname, 
                'price': plPrice,
                'nation': plCountry,
                'club': plClub,
                'pos': plPos,
                'rate': plRating,
                'foot': plFoot,
                'height': plHeight,
            },
            'stats': {
                'speed': plSpeed,
                'acc': plAcc,
                'stamina': plStamina,
                'control': plControl,
                'stength': plStrength,
                'tackle': plTackle,
                'passing': plPassing,
                'shooting': plShooting,
                # GK Stats
                'gk_handling': plGKHandling,
                'gk_reaction': plGKReaction
            }
        }
    return playerDict

legPlayers, rarePlayers = playerConverter('Legendary Players', 'spreadsheets\DLS-23-RATINGS.xlsx'), playerConverter('Rare Players', 'spreadsheets\DLS-23-RATINGS.xlsx')
playersDict = {**legPlayers, **rarePlayers}

with open('json\PlayerInfo.json', 'w')as f:
    json_string = json.dumps(playersDict, indent=3)
    f.write(json_string)

def getPlayers():
    legPlayers, rarePlayers = playerConverter('Legendary Players', 'spreadsheets\DLS-23-RATINGS.xlsx'), playerConverter('Rare Players', 'spreadsheets\DLS-23-RATINGS.xlsx')
    return {'legends': legPlayers, 'rare': rarePlayers}